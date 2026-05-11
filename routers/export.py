from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from datetime import datetime, date, timedelta
from typing import Optional
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

import models
from auth import get_current_user
from database import get_db

router = APIRouter(prefix="/export", tags=["export"])

_HEADER_FILL = PatternFill("solid", fgColor="2E7D32")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_ALT_FILL = PatternFill("solid", fgColor="F1F8E9")
_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)


def _style_header(ws, row: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _BORDER


def _style_row(ws, row: int, col_count: int, alt: bool = False):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row, column=col)
        if alt:
            cell.fill = _ALT_FILL
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _BORDER


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max(max_len + 2, 10), 40)


def _get_range(period: str, date_from: Optional[str], date_to: Optional[str]):
    today = date.today()
    if period == "today":
        return datetime.combine(today, datetime.min.time()), datetime.utcnow()
    if period == "week":
        return datetime.combine(today - timedelta(days=7), datetime.min.time()), datetime.utcnow()
    if period == "month":
        return datetime.combine(today - timedelta(days=30), datetime.min.time()), datetime.utcnow()
    if period == "custom" and date_from and date_to:
        return (
            datetime.strptime(date_from, "%Y-%m-%d"),
            datetime.strptime(date_to, "%Y-%m-%d").replace(hour=23, minute=59),
        )
    return datetime.combine(today, datetime.min.time()), datetime.utcnow()


@router.get("/batches")
def export_batches(
    period: str = "today",
    site_id: Optional[int] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(get_current_user),
):
    start, end = _get_range(period, date_from, date_to)

    resolved_site_id = site_id
    if not resolved_site_id and current_user.role in ("operator", "master"):
        resolved_site_id = current_user.site_id

    query = db.query(models.WasteBatch).filter(
        models.WasteBatch.created_at >= start,
        models.WasteBatch.created_at <= end,
    )
    if resolved_site_id:
        query = query.filter(models.WasteBatch.site_id == resolved_site_id)
    batches = query.order_by(models.WasteBatch.created_at.desc()).all()

    deviations_q = db.query(models.Deviation).join(models.WasteBatch).filter(
        models.WasteBatch.created_at >= start,
        models.WasteBatch.created_at <= end,
    )
    if resolved_site_id:
        deviations_q = deviations_q.filter(models.WasteBatch.site_id == resolved_site_id)
    all_deviations = deviations_q.all()

    wb = openpyxl.Workbook()

    # --- Sheet 1: Summary ---
    ws_sum = wb.active
    ws_sum.title = "Сводка"
    ws_sum.column_dimensions["A"].width = 30
    ws_sum.column_dimensions["B"].width = 20

    info_rows = [
        ("Отчёт сформирован", datetime.utcnow().strftime("%d.%m.%Y %H:%M")),
        ("Период с", start.strftime("%d.%m.%Y")),
        ("Период по", end.strftime("%d.%m.%Y")),
        ("Объект", db.query(models.Site).filter(models.Site.id == resolved_site_id).first().name
         if resolved_site_id else "Все объекты"),
        ("Всего партий", len(batches)),
        ("Завершённых партий", sum(
            1 for b in batches
            if b.stages and all(s.status == "completed" for s in b.stages)
        )),
        ("Всего отклонений", len(all_deviations)),
    ]

    for r_idx, (label, value) in enumerate(info_rows, start=1):
        ws_sum.cell(r_idx, 1, label).font = Font(bold=True)
        ws_sum.cell(r_idx, 2, str(value))

    # KPI block
    total_stages = sum(
        1 for b in batches for s in b.stages
        if s.status == "completed" and s.started_at and s.completed_at
    )
    on_time = sum(
        1 for b in batches for s in b.stages
        if s.status == "completed" and s.started_at and s.completed_at
        and (s.completed_at - s.started_at).seconds // 60 <= s.norm_minutes
    )
    pct = round(on_time / total_stages * 100, 1) if total_stages > 0 else 100.0

    ws_sum.cell(len(info_rows) + 2, 1, "% выполнения в норматив").font = Font(bold=True)
    ws_sum.cell(len(info_rows) + 2, 2, f"{pct}%")

    # --- Sheet 2: Batches ---
    ws_b = wb.create_sheet("Партии")
    batch_headers = [
        "ID", "Название отходов", "Код ФККО", "Класс опасности",
        "Объём", "Объект", "Оператор", "Статус партии",
        "Текущий этап", "Дата создания",
    ]
    for col, h in enumerate(batch_headers, start=1):
        ws_b.cell(1, col, h)
    _style_header(ws_b, 1, len(batch_headers))
    ws_b.row_dimensions[1].height = 30

    for row_idx, b in enumerate(batches, start=2):
        all_done = b.stages and all(s.status == "completed" for s in b.stages)
        stage = b.current_stage
        values = [
            b.id,
            b.waste_name,
            b.fkko_code,
            b.hazard_class,
            b.volume,
            b.site.name if b.site else "",
            b.operator.full_name if b.operator else "",
            "Завершена" if all_done else "Активна",
            stage.stage_name if stage else "",
            b.created_at.strftime("%d.%m.%Y %H:%M") if b.created_at else "",
        ]
        for col, v in enumerate(values, start=1):
            ws_b.cell(row_idx, col, v)
        _style_row(ws_b, row_idx, len(batch_headers), alt=(row_idx % 2 == 0))
    _auto_width(ws_b)

    # --- Sheet 3: Stages analytics ---
    ws_st = wb.create_sheet("Этапы")
    stage_headers = [
        "Партия ID", "Название партии", "Объект",
        "Этап", "Норматив (мин)", "Факт (мин)", "Вовремя",
        "Начало", "Завершение",
    ]
    for col, h in enumerate(stage_headers, start=1):
        ws_st.cell(1, col, h)
    _style_header(ws_st, 1, len(stage_headers))
    ws_st.row_dimensions[1].height = 30

    st_row = 2
    for b in batches:
        for s in b.stages:
            if s.status == "completed" and s.started_at and s.completed_at:
                duration = (s.completed_at - s.started_at).seconds // 60
                in_norm = "Да" if duration <= s.norm_minutes else "Нет"
                values = [
                    b.id, b.waste_name,
                    b.site.name if b.site else "",
                    s.stage_name, s.norm_minutes, duration, in_norm,
                    s.started_at.strftime("%d.%m.%Y %H:%M"),
                    s.completed_at.strftime("%d.%m.%Y %H:%M"),
                ]
                for col, v in enumerate(values, start=1):
                    ws_st.cell(st_row, col, v)
                _style_row(ws_st, st_row, len(stage_headers), alt=(st_row % 2 == 0))
                st_row += 1
    _auto_width(ws_st)

    # --- Sheet 4: Deviations ---
    ws_d = wb.create_sheet("Отклонения")
    dev_headers = [
        "ID", "Партия", "Объект", "Тип отклонения",
        "Описание", "Есть фото", "Дата",
    ]
    for col, h in enumerate(dev_headers, start=1):
        ws_d.cell(1, col, h)
    _style_header(ws_d, 1, len(dev_headers))
    ws_d.row_dimensions[1].height = 30

    for row_idx, d in enumerate(all_deviations, start=2):
        values = [
            d.id, d.batch.waste_name,
            d.batch.site.name if d.batch.site else "",
            d.type, d.description,
            "Да" if d.photo_path else "Нет",
            d.created_at.strftime("%d.%m.%Y %H:%M") if d.created_at else "",
        ]
        for col, v in enumerate(values, start=1):
            ws_d.cell(row_idx, col, v)
        _style_row(ws_d, row_idx, len(dev_headers), alt=(row_idx % 2 == 0))
    _auto_width(ws_d)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"waste_report_{start.strftime('%Y%m%d')}_{end.strftime('%Y%m%d')}.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
